import time
import openpyxl
import re
import os

from typing import List, Dict, Optional
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# URL страницы
SITE = 'https://online.metro-cc.ru/category'
CATEGORY = '/myasnye'
SUBCATEGORY = '/myaso'

def initialize_driver() -> webdriver.Chrome:
    """Инициализация драйвера."""
    driver = webdriver.Chrome()
    driver.maximize_window()
    return driver

def open_page(driver: webdriver.Chrome, url: str) -> None:
    """Открытие страницы."""
    driver.get(url)
    time.sleep(5)  # Ожидание загрузки страницы

def click_filter(driver: webdriver.Chrome) -> None:
    """Клик на элемент в боковом меню."""
    try:
        filter = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'catalog-checkbox__icon')]"))
        )
        filter.click()
        time.sleep(3)  # Ожидание загрузки новых данных
    except NoSuchElementException:
        print("Элемент в боковом меню не найден")

def load_products(driver: webdriver.Chrome, desired_product_count: int, timeout: int) -> BeautifulSoup:
    """Скроллинг страницы и нажатие на кнопку 'Показать ещё' до тех пор, пока не будет загружено нужное количество товаров или не истечет время."""
    start_time = time.time()
    end_time = start_time + timeout

    while time.time() < end_time:
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        products = soup.find_all('div', class_='product-card__content')
        current_product_count = len(products)

        if current_product_count >= desired_product_count:
            break

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        try:
            show_more_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'subcategory-or-type__load-more') and contains(., 'Показать ещё')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", show_more_button)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", show_more_button)
            time.sleep(2)
        except NoSuchElementException:
            break
        except ElementClickInterceptedException:
            driver.execute_script("window.scrollBy(0, 200);")
            time.sleep(1)

    return soup

def parse_product_data(driver: webdriver.Chrome, soup: BeautifulSoup, desired_product_count: int) -> List[Dict[str, Optional[str]]]:
    """Парсинг данных о товарах."""
    products = soup.find_all('div', class_='catalog-2-level-product-card')
    product_data = []

    for product in products[:desired_product_count]:
        name_element = product.find('span', class_="product-card-name__text")
        price_element = product.find('span', class_="product-price nowrap product-unit-prices__actual style--catalog-2-level-product-card-major-actual color--red")
        link_element = product.find('a', class_="product-card-name reset-link catalog-2-level-product-card__name style--catalog-2-level-product-card")
        past_price_element = product.find('span', class_="product-price nowrap product-unit-prices__old style--catalog-2-level-product-card-major-old")

        if name_element and price_element and link_element:
            name = re.sub(r',.*', '', name_element.text.strip())
            promo_price = re.sub(r'\D', '', price_element.text.strip())
            product_id = product.get('data-sku')
            product_link = 'https://online.metro-cc.ru' + link_element.get('href')
            past_price = re.sub(r'\D', '', past_price_element.text.strip()) if past_price_element else None

            driver.get(product_link)
            time.sleep(2)
            product_page_html = driver.page_source
            product_page_soup = BeautifulSoup(product_page_html, 'html.parser')
            brand_element = product_page_soup.find('li', class_="product-attributes__list-item")
            brand = brand_element.find('a').text.strip() if brand_element else None

            product_data.append({
                'id': product_id,
                'name': name,
                'link': product_link,
                'past_price': past_price,
                'promo_price': promo_price,
                'brand': brand
            })

    return product_data

def save_to_excel(product_data: List[Dict[str, Optional[str]]], filename: str, city: str) -> None:
    """Сохранение данных в Excel."""
    # Создаем папку с именем города, если она не существует
    city_folder = os.path.join(os.getcwd(), city)
    if not os.path.exists(city_folder):
        os.makedirs(city_folder)

    # Формируем полный путь к файлу
    full_filename = os.path.join(city_folder, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    headers = ['ID товара', 'Наименование', 'Ссылка на товар', 'Регулярная цена', 'Промо цена', 'Бренд']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, data in enumerate(product_data, 2):
        ws.cell(row=row_num, column=1, value=data['id'])
        ws.cell(row=row_num, column=2, value=data['name'])
        ws.cell(row=row_num, column=3, value=data['link'])
        ws.cell(row=row_num, column=4, value=data['past_price'])
        ws.cell(row=row_num, column=5, value=data['promo_price'])
        ws.cell(row=row_num, column=6, value=data['brand'])

    wb.save(full_filename)

def select_city(driver: webdriver.Chrome, city: str,adress_count: int) -> None:

    """Выбор города (Москва или Санкт-Петербург)."""
    # Нажимаем на кнопку с адресом
    address_button = WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable((By.CLASS_NAME, "header-address__receive-button"))
    )
    print("Нажимаем на кнопку с адресом")
    driver.execute_script("arguments[0].click();", address_button)
    time.sleep(2)

    # Нажимаем на кнопку "Самовывоз" внутри <div class="delivery__tab">
    pickup_button = WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable((By.XPATH, "//div[@class='delivery__tab']"))
    )
    driver.execute_script("arguments[0].click();", pickup_button)
    time.sleep(2)
    print("Нажимаем на кнопку 'Самовывоз'")


    if city == 'Москва':
        if adress_count == 0:
            apply_button = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "delivery__btn-apply"))
            )
            driver.execute_script("arguments[0].click();", apply_button)
            time.sleep(2)
            print("Нажимаем на кнопку 'Выбрать'")
        elif adress_count >= 1:

            address_radios = WebDriverWait(driver, 2).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "pickup-form__address-radio"))
            )
            for address_radio in address_radios:
                driver.execute_script("arguments[0].click();", address_radios[adress_count])
                time.sleep(1)
                print(f"Выбираем адрес {adress_count}")
                apply_button = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "delivery__btn-apply"))
                )
                driver.execute_script("arguments[0].click();", apply_button)
                time.sleep(2)
                print("Нажимаем на кнопку 'Выбрать'")
                return
            else:
                print(f"Адрес {adress_count} не найден")

            return 
   
        adress_count += 1
        return

    
    # Нажимаем на кнопку "Изменить"
    change_button = WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'Изменить')]"))
    )
    driver.execute_script("arguments[0].click();", change_button)
    time.sleep(2)
    print("Нажимаем на кнопку 'Изменить'")




    # Выбираем нужный город
    city_button = WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '{city}')]".format(city=city)))
    )
    driver.execute_script("arguments[0].click();", city_button)
    time.sleep(2)
    print("Выбираем город")

    
    address_radios = WebDriverWait(driver, 2).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "pickup-form__address-radio"))
    )
    if adress_count < len(address_radios):
        driver.execute_script("arguments[0].click();", address_radios[adress_count])
        time.sleep(1)
        print(f"Выбираем адрес {adress_count}")
    else:
        print(f"Адрес {adress_count} не найден")


    # Нажимаем на кнопку "Применить"
    apply_button = WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable((By.CLASS_NAME, "delivery__btn-apply"))
    )
    driver.execute_script("arguments[0].click();", apply_button)
    time.sleep(2)
    print("Нажимаем на кнопку 'Применить'")

def main() -> None:
    adress_count = 0
    url = SITE + CATEGORY + SUBCATEGORY
    desired_product_count = 10
    timeout = 100


    city = "Санкт-Петербург"
    while adress_count < 3:
        print(f"Начало цикла {adress_count} для города {city}")
        driver = initialize_driver()
        if driver:
            open_page(driver, url)
            select_city(driver, city, adress_count)
            click_filter(driver)
            soup = load_products(driver, desired_product_count, timeout)
            product_data = parse_product_data(driver, soup, desired_product_count)
            driver.quit()
            save_to_excel(product_data, f"products{adress_count}{city}.xlsx",city=city)
            adress_count += 1
        else:
            print("Не удалось инициализировать драйвер.")
            break


    adress_count = 0
    city = "Москва"
    while adress_count < 11:
        print(f"Начало цикла {adress_count} для города {city}")
        driver = initialize_driver()
        if driver:
            open_page(driver, url)
            select_city(driver, city, adress_count)
            click_filter(driver)
            soup = load_products(driver, desired_product_count, timeout)
            product_data = parse_product_data(driver, soup, desired_product_count)
            driver.quit()
            save_to_excel(product_data, f"products{adress_count}{city}.xlsx",city=city)
            adress_count += 1
        else:
            print("Не удалось инициализировать драйвер.")
            break

if __name__ == "__main__":
    main()