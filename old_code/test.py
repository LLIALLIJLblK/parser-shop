import time
import openpyxl
import re

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# URL страницы
SITE = 'https://online.metro-cc.ru/category'
CATEGORY = '/myasnye'
SUBCATEGORY = '/myaso'

url = (
    SITE + CATEGORY + SUBCATEGORY    
)

# Инициализация драйвера
driver = webdriver.Chrome()
driver.maximize_window()

# Открытие страницы
driver.get(url)

# Ожидание загрузки страницы
time.sleep(5)

# Клик на элемент в боковом меню
#1
try:
    filter = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'catalog-checkbox__icon')]"))
    )
    filter.click()
    time.sleep(3)  # Ожидание загрузки новых данных
except NoSuchElementException:
    print("Элемент в боковом меню не найден")

# Желаемое количество товаров
desired_product_count = 5

# Таймер на 30 секунд
start_time = time.time()
end_time = start_time + 30

# Скроллинг страницы и нажатие на кнопку "Показать ещё" до тех пор, пока не будет загружено нужное количество товаров или не истечет время
while time.time() < end_time:
    # Получение HTML-кода страницы
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    
    # Подсчет текущего количества товаров
    products = soup.find_all('div', class_='product-card__content')
    current_product_count = len(products)
    
    # Если достигнуто желаемое количество товаров, выходим из цикла
    if current_product_count >= desired_product_count:
        break
    
    # Скроллинг страницы до конца
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)
    
    # Поиск и нажатие на кнопку "Показать ещё"
    try:
        show_more_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'subcategory-or-type__load-more') and contains(., 'Показать ещё')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", show_more_button)  # Прокрутка к кнопке
        time.sleep(1)  # Ожидание прокрутки
        driver.execute_script("arguments[0].click();", show_more_button)  # Нажатие на кнопку с помощью JavaScript
        time.sleep(2)  # Ожидание загрузки новых товаров
    except NoSuchElementException:
        # Если кнопка "Показать ещё" не найдена, выходим из цикла
        break
    except ElementClickInterceptedException:
        # Если кнопка перекрыта другим элементом, прокручиваем страницу и повторяем попытку
        driver.execute_script("window.scrollBy(0, 200);")
        time.sleep(1)

# Парсинг HTML с помощью BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

# Пример: вывод всех данных о товарах
products = soup.find_all('div', class_='catalog-2-level-product-card')

product_data = []

for product in products[:desired_product_count]:
    name_element = product.find('span', class_="product-card-name__text")
    price_element = product.find('span', class_="product-price nowrap product-unit-prices__actual style--catalog-2-level-product-card-major-actual color--red")
    link_element = product.find('a', class_="product-card-name reset-link catalog-2-level-product-card__name style--catalog-2-level-product-card")
    past_price_element = product.find('span', class_="product-price nowrap product-unit-prices__old style--catalog-2-level-product-card-major-old")
    
    if name_element and price_element and link_element:
        name = name_element.text.strip()
        name = re.sub(r',.*', '', name)
        
        promo_price = price_element.text.strip()
        promo_price = re.sub(r'\D', '', promo_price)
        
        product_id = product.get('data-sku')
        product_link = 'https://online.metro-cc.ru' + link_element.get('href')
        past_price = past_price_element.text.strip() if past_price_element else None
        past_price = re.sub(r'\D', '', past_price) if past_price else None
        
        # Переходим на страницу товара для получения бренда
        driver.get(product_link)
        time.sleep(2)  # Ожидание загрузки страницы товара
        
        product_page_html = driver.page_source
        product_page_soup = BeautifulSoup(product_page_html, 'html.parser')
        
        # Ищем элемент с брендом
        brand_element = product_page_soup.find('li', class_="product-attributes__list-item")
        brand = brand_element.find('a').text.strip() if brand_element else None
        
        product_data.append({
            'id': product_id,
            'name': name,
            'link': product_link,
            'past_price': past_price,
            'promo_price': promo_price,
            'brand': brand
        })

# Закрытие драйвера
driver.quit()

# Сохранение данных в Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Товары"

# Заголовки
headers = ['ID товара', 'Наименование', 'Ссылка на товар', 'Регулярная цена', 'Промо цена', 'Бренд']
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Заполнение данных
for row_num, data in enumerate(product_data, 2):
    ws.cell(row=row_num, column=1, value=data['id'])
    ws.cell(row=row_num, column=2, value=data['name'])
    ws.cell(row=row_num, column=3, value=data['link'])
    ws.cell(row=row_num, column=4, value=data['past_price'])
    ws.cell(row=row_num, column=5, value=data['promo_price'])
    ws.cell(row=row_num, column=6, value=data['brand'])

# Сохранение файла
wb.save("products.xlsx")